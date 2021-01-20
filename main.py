from tkinter import *
from tkinter import ttk
from tkinter import filedialog
from tkinter import Listbox
import openpyxl


def fileOpen():
    filename.set(filedialog.askopenfilename(initialdir="", title="choose excel file", filetypes=(("excel files", "*.xlsx"), ("all files", "*.*"))))


def getSheetName():
    book = openpyxl.load_workbook(filename.get())
    sheetname = book.sheetnames
    combobox['values'] = ([sheet for sheet in sheetname])


def readFile():
    global book
    book = openpyxl.load_workbook(filename.get())
    name = combobox.get()
    sheet = book[name]

    # data read
    for i in range(sheet.max_row):
        test_id[i] = sheet['A'][i].value
        test_id_year[i] = sheet['B'][i].value
        test_id_month[i] = sheet['C'][i].value
        test_id_day[i] = sheet['D'][i].value
        test_id_building[i] = sheet['E'][i].value
        test_id_room[i] = sheet['F'][i].value
        test_id_desc[i] = 'blur'
        # print(test_id[i], test_id_year[i], test_id_month[i], test_id_day[i])

        listbox.insert(i, test_id[i])

    # display
    # for i in range(len(test_id)):
    listbox.pack()


def writeFile():
    name = combobox.get()
    sheet = book[name]

    for i in range(len(test_id)):
        sheet.cell(row=i+1, column=1, value=test_id[i])
        sheet.cell(row=i+1, column=2, value=test_id_year[i])
        sheet.cell(row=i+1, column=3, value=test_id_month[i])
        sheet.cell(row=i+1, column=4, value=test_id_day[i])
        sheet.cell(row=i+1, column=5, value=test_id_desc[i])
        sheet.cell(row=i+1, column=6, value=test_id_result[i])
        sheet.cell(row=i+1, column=7, value=test_id_memo[i])
        
    book.save(filename.get())

if __name__ == "__main__":
    # main window creation
    root = Tk()
    root.title("Maintain Helper")
    root.geometry("640x520+100+100")
    root.resizable(False, False)

    # variables
    filename = StringVar()
    sheetname = StringVar()
    test_id_building = {}
    test_id_room = {}
    test_id = {}
    test_id_year = {}
    test_id_month = {}
    test_id_day = {}
    test_id_desc = {}
    test_id_result = {}
    test_id_memo = {}

    lbl1 = Label(root, text="file name: ")
    lbl1.place(x=10, y=10)
    # lbl.pack()

    txt = Entry(root, textvariable=filename, width=58)
    txt.place(x=100, y=10)
    # txt.pack()

    # get file name
    btn1 = Button(root, text=" . . . ", command=fileOpen)
    btn1.place(x=580, y=10)
    # btn.pack(side=RIGHT)

    lbl2 = Label(root, text="sheet name: ")
    lbl2.place(x=10, y=40)

    # sheet select
    combobox = ttk.Combobox(root, width=56, textvariable=sheetname, postcommand=getSheetName)
    combobox.place(x=100, y=40)

    # data display frame
    listbox = Listbox(root, selectmode='extended', height=0)
    listbox.yview()
    listbox.place(x=10, y=110)
    
    # read file
    btn2 = Button(root, text="  read file  ", command=readFile)
    btn2.place(x=540, y=40)

    # write file
    btn3 = Button(root, text="  write file  ", command=writeFile)
    btn3.place(x=540, y=70)

    # test list display frame 
    #

    # main loop
    root.mainloop()
